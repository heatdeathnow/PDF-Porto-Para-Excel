from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import get_column_letter
from multiprocessing import cpu_count
from threading import current_thread
from utils import to_num, get_sub
from colorama import Fore, Style
from pdfplumber.page import Page
from time import perf_counter
from pandas import DataFrame
from vars import max_threads
from threading import Lock
import pandas as pd
from crop import *


def get_cells(page: Page, rel_pos: list[int]) -> list:
    lines = page.extract_text_lines(return_chars = False)
    cells = [''] * len(rel_pos)
    flag = False

    for i, line in enumerate(lines):
        if i + 1 == len(lines) and not flag:  # Se é o último.
            index = rel_pos.index(round(line['top']))
            cells[index] = line['text']

        elif flag:  # Isso basicamente tem o efeito de pular um round da iteração.
            flag = False
        
        elif abs(line['top'] - lines[i + 1]['top']) < 10:  # Se estiver muito pergunto verticalmente, considerar como duas linhas do mesmo texto
            index = rel_pos.index(round(line['top']))
            cells[index] = '\n'.join([line['text'], lines[i + 1]['text']])
            flag = True

        elif not flag:
            index = rel_pos.index(round(line['top']))
            cells[index] = line['text']

    return cells

def get_relative_positions(page: Page) -> list:  # Passar idade pois é mais fácil
    positions = []
    lines = page.extract_text_lines(return_chars = False)

    for line in lines:
        positions.append(round(line['top']))

    return positions

def get_info_from_leftside(page: Page) -> DataFrame:
    ci = crop_idade(page)
    rp = get_relative_positions(ci)
    idade = get_cells(ci, rp)

    sub = get_sub(page)

    with ThreadPoolExecutor(max_workers = max_threads / max_threads) as tpx:
        cropped_pages = list(tpx.map(crop_several, [page] * 8, [0, 1, 2, 3, 5, 6, 7, 8]))
        cols = list(tpx.map(get_cells, cropped_pages, [rp] * 8))

    df = DataFrame({'Subestipulante': [sub] * len(idade),
                    'Seguro': cols[0],
                    'Dep': cols[1],
                    'Nome do Segurado': cols[2],
                    'Reg. Func.': cols[3],
                    'Idade': idade,
                    'Parentesco': cols[4],
                    'Plano': cols[5],
                    'Mov./Inic.Vig.': cols[6],
                    'Situação do Segurado': cols[7]})
    
    return df

def get_info_from_rightside(page: Page) -> DataFrame:
    rightside = crop_right_side(page)

    # Isso é necessário porque as vezes os faturamentos da porto vêm faltando o prêmio. É preciso saber o tamanho de antemão para evitar problemas.
    size = max(rightside.extract_text().count('Prêmio Base'), rightside.extract_text().count('Desc por Co-Part'))

    premio = [''] * size
    rataq = [''] * size
    ratav = [''] * size
    consultasq = [''] * size
    consultasv = [''] * size
    examesq = [''] * size
    examesv = [''] * size
    socorroq =[''] * size
    socorrov = [''] * size
    copart = [''] * size
    bonificacao = [''] * size
    acerto = [''] * size
    iof = [''] * size

    i = -1  # Prêmio Base vai ser sempre o primeiro valor.
    for line in rightside.extract_text_lines(return_chars = False):
        match line['text']:
            case x if 'Prêmio Base' in x:
                premio[i + 1] = line['text'][12:].strip()
                i += 1

            case x if 'CONSULTAS' in x:
                end = line['text'].index(')')
                consultasq[i] = line['text'][11:end]  # O "(" sempre estará na 11ª posição.
                consultasv[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

            case x if 'EXAMES' in x:
                end = line['text'].index(')')
                examesq[i] = line['text'][8:end]  # O "(" sempre estará na 8ª posição.
                examesv[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

            case x if 'PRONTO-SOCORRO' in x:
                end = line['text'].index(')')
                socorroq[i] = line['text'][16:end]  # O "(" sempre estará na 16ª posição.
                socorrov[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

            case x if 'Pro-Rata' in x:
                end = line['text'].index(')')
                end_ = line['text'].index(')', end + 1)
                rataq[i] = line['text'][10:end]  # O "(" sempre estará na 16ª posição.
                ratav[i] = line['text'][end_ + 1:].strip()  # Começando depois do ")" retire todos os espaços.

            case x if 'Desc por Co-Part' in x:
                if copart[i] != '':  # Pode acontecer de faltar o prêmio por erro da Porto Seguro.
                    i += 1
                copart[i] = line['text'][22:].strip()

            case x if 'Bonificação GRES(-)' in x:
                bonificacao[i] = line['text'][19:].strip()
        
            case x if 'Acerto (-)' in x:
                acerto[i] = line['text'][10:].strip()
            
            case x if 'IOF' in x:
                if iof[i] != '' and iof[i] != '0,00' and iof[i] != 0:
                    pass  # Por algum motivo, os PDFs da Porto podem vir com dois IOFs, o segundo ou o primeiro erroneamente zerados.
                else:
                    iof[i] = line['text'][4:].strip()

            case x if 'TOTAL' not in x: 
                print(f"{Fore.RED}ERRO: Não foi possível ler a linha: {line['text']}{Style.RESET_ALL}")
                pass

    df = DataFrame({'Prêmio Base': premio,
                    'Pro-Rata (quantidade)': rataq,
                    'Pro-Rata (valor)': ratav,
                    'CONSULTAS (quantidade)': consultasq,
                    'CONSULTAS (valor)': consultasv,
                    'EXAMES (quantidade)': examesq,
                    'EXAMES (valor)': examesv,
                    'PRONTO-SOCORRO (quantidade)': socorroq,
                    'PRONTO-SOCORRO (valor)': socorrov,
                    'Desc por Co-Part. (-)': copart,
                    'Bonificação GRES(-)': bonificacao,
                    'Acerto (-)': acerto,
                    'IOF': iof})
    
    return df

def get_info_from_qe(page: Page) -> DataFrame:
    qev = crop_qe_valores(page)
    cells = {}

    for i, row in enumerate(qev.extract_text_lines()):
        if i == 1: continue

        for j, value in enumerate(row['text'].split(' ')):
            cells[f'{get_column_letter(j + 2)}{i + 2}'] = to_num(value)  # B2:G2 & B4:G6

    return cells

def get_info_from_premio(page: Page) -> DataFrame:
    premio = crop_premio(page)
    cells = {}

    for i, row in enumerate(premio.extract_text_lines()):
        if i == 3: continue

        for j, value in enumerate(row['text'].split(' ')):
            cells[f'{get_column_letter(j + 2)}{i + 2}'] = to_num(value)

    return cells

def get_info_from_totalizador(page: Page) -> DataFrame:
    total = crop_totalizador(page)
    cells = {}

    for i, row in enumerate(total.extract_text_lines()):
        j = row['text'].index(':') + 2
        cells[f'B{i + 1}'] = to_num(row['text'][j:])

    return cells

def get_info_from_detalhamento(page: Page) -> DataFrame:
    detalhe = crop_detalhamento(page)
    cells = {}

    for row in detalhe.extract_text_lines():
        if 'Em atendimento à Lei' in row['text']: break  # É assim que se sabe quando acabou a informação útil.

        l = 0
        for i in 'qwertyuiopasdfghjklçzxcvbnm':  # Isso descobre a posição da última *letra* na linha.
            j = row['text'].lower().rfind(i)
            if j > l: l = j

        hold = []
        for i, cell in enumerate(row['text'][l + 2:].split(' ')):
            if i == 3: continue
            hold.append(to_num(cell))

        cells[row['text'][:l + 1]] = hold
    
    return cells

def get_info_from_page(page: Page, lock: Lock) -> DataFrame:
    start_time = perf_counter()
    check = crop_headers(page)
    
    # Por algum motivo, mesmo nada sendo compartilhado, isso deve ser feito com um lock senão o PDFPlumber dá erro.
    # Eu tenho bastante certeza que duas threads olhando a mesma página ao mesmo tempo gera erros, mas esse não é o caso aqui.
    with lock:  
        if len(check.extract_words()) == 0:
            is_empty = True
        else:
            is_empty = False

    if is_empty:
        return DataFrame({'Subestipulante': [],
                          'Seguro': [],
                          'Dep': [],
                          'Nome do Segurado': [],
                          'Reg. Func.': [],
                          'Idade': [],
                          'Parentesco': [],
                          'Plano': [],
                          'Mov./Inic.Vig.': [],
                          'Situação do Segurado': [],
                          'Prêmio Base': [],
                          'Pro-Rata (quantidade)': [],
                          'Pro-Rata (valor)': [],
                          'CONSULTAS (quantidade)': [],
                          'CONSULTAS (valor)': [],
                          'EXAMES (quantidade)': [],
                          'EXAMES (valor)': [],
                          'PRONTO-SOCORRO (quantidade)': [],
                          'PRONTO-SOCORRO (valor)': [],
                          'Desc por Co-Part. (-)': [],
                          'Bonificação GRES(-)': [],
                          'Acerto (-)': [],
                          'IOF': []})
    
    else:
        dfl = get_info_from_leftside(page)
        dfr = get_info_from_rightside(page)
        print(f'{current_thread().name}: informação da {page.page_number:02}ª página lida em {perf_counter() - start_time:.2f} segundos.')
        return pd.concat([dfl, dfr], axis = 'columns')
