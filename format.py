from openpyxl.styles import PatternFill, Color, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from vars import *
import vars


def format_sheets(wb: Workbook) -> None:
    blue = Color('FF002060')
    white = Color('FFFFFFFF')
    black = Color('FF000000')
    side = Side(color = black, style = 'thin')
    borders = Border(side, side, side, side)
    blue_fill = PatternFill(start_color = blue, end_color = blue, fill_type = 'solid')  # Não sei por que bgcolor não funciona aqui.
    header_font = Font(color = white, bold = True)

    ws = wb['Informação extraída']
    ws.sheet_view.showGridLines = False
    for i in range(1, ws.max_column + 1):
        if get_column_letter(i) in (DEP, REG, SIT, COQ, COV, EXQ, EXV, PRQ, PRV, ):    # Oculta essas colunas
            ws.column_dimensions[f'{get_column_letter(i)}'].hidden = True

        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, ws.max_row + 1):
            ws[f'{PRE}{j}'].number_format = nformat
            ws[f'{RAV}{j}'].number_format = nformat
            ws[f'{COV}{j}'].number_format = nformat
            ws[f'{EXV}{j}'].number_format = nformat
            ws[f'{PRV}{j}'].number_format = nformat
            ws[f'{DES}{j}'].number_format = nformat
            ws[f'{BON}{j}'].number_format = nformat
            ws[f'{ACE}{j}'].number_format = nformat
            ws[f'{COP}{j}'].number_format = nformat
            ws[f'{TDE}{j}'].number_format = nformat
            ws[f'{PRR}{j}'].number_format = nformat
            ws[f'{get_column_letter(i)}{j}'].border = borders
    
    ws = wb['Quantidades e valores']
    ws.sheet_view.showGridLines = False
    for i in range(1, 8):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, 7):
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font

            ws[f'{get_column_letter(i)}{j}'].border = borders
            ws[f'C{j}'].number_format = nformat
            ws[f'E{j}'].number_format = nformat
            ws[f'G{j}'].number_format = nformat
    
    ws[f'I3'].number_format = nformat
    ws[f'J3'].number_format = nformat
    ws[f'I3'].border = borders
    ws[f'J3'].border = borders
    ws[f'I1'].border = borders
    ws[f'I2'].border = borders
    ws[f'J2'].border = borders
    ws[f'I2'].fill = blue_fill
    ws[f'I1'].fill = blue_fill
    ws[f'J2'].fill = blue_fill
    ws[f'I2'].font = header_font
    ws[f'I1'].font = header_font
    ws[f'J2'].font = header_font
    ws.merge_cells('I1:J1')

    ws = wb['Prêmio']
    ws.sheet_view.showGridLines = False
    for i in range(1, 5):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, 7):
            if i == 1:
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font

            else: 
                ws[f'{get_column_letter(i)}{j}'].number_format = nformat

            ws[f'{get_column_letter(i)}{j}'].border = borders
            
    ws = wb['Totalizador']
    ws.sheet_view.showGridLines = False
    for i in range(1, 3):
        if i != 2:
            ws[f'{get_column_letter(i)}1'].fill = blue_fill
            ws[f'{get_column_letter(i)}1'].font = header_font

        else:
            ws[f'{get_column_letter(i)}1'].number_format = nformat

        ws[f'{get_column_letter(i)}1'].border = borders
        for j in range(2, 9):
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font

            else: 
                ws[f'{get_column_letter(i)}{j}'].number_format = nformat
            
            ws[f'{get_column_letter(i)}{j}'].border = borders

    ws = wb['Detalhamento']
    ws.sheet_view.showGridLines = False
    for i in range(1, 7):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, ws.max_row):
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font

            elif i != 2: 
                ws[f'{get_column_letter(i)}{j}'].number_format = nformat
            
            ws[f'{get_column_letter(i)}{j}'].border = borders

    ws[f'E{ws.max_row}'].number_format = nformat
    ws[f'F{ws.max_row}'].number_format = nformat
    ws[f'E{ws.max_row}'].border = borders
    ws[f'F{ws.max_row}'].border = borders

def check_values(wb: Workbook) -> None:
    red = Color('00FF0000')
    green = Color('00008000')
    wrong = PatternFill(bgColor = red, fill_type = 'solid')
    right = PatternFill(bgColor = green, fill_type = 'solid')

    ws = wb['Quantidades e valores']
    for key in vars.qev.keys():
        ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [vars.qev[key]], fill = right))
        ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [vars.qev[key]], fill = wrong))

    ws.conditional_formatting.add('I3', CellIsRule(operator = 'equal', formula = [vars.qev['C4']], fill = right))
    ws.conditional_formatting.add('I3', CellIsRule(operator = 'notEqual', formula = [vars.qev['C4']], fill = wrong))
    ws.conditional_formatting.add('J3', CellIsRule(operator = 'equal', formula = [vars.qev['E4']], fill = right))
    ws.conditional_formatting.add('J3', CellIsRule(operator = 'notEqual', formula = [vars.qev['E4']], fill = wrong))

    ws = wb['Prêmio']
    for key in vars.premio.keys():
        if '4' not in key:
            ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [vars.premio[key]], fill = right))
            ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [vars.premio[key]], fill = wrong))

        else:
            ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [-vars.premio[key]], fill = right))
            ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [-vars.premio[key]], fill = wrong))

    ws = wb['Totalizador']
    for key in vars.total.keys():
        ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [vars.total[key]], fill = right))
        ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [vars.total[key]], fill = wrong))
    
    ws = wb['Detalhamento']
    for i in range(2, ws.max_row):
        for j in range(2, 6):
            key = str(ws[f'A{i}'].value)
            key = key.replace('\n', ' ')
            ws.conditional_formatting.add(f'{get_column_letter(j)}{i}', CellIsRule(operator = 'equal', formula = [vars.detalhe[key][j - 2]], fill = right))
            ws.conditional_formatting.add(f'{get_column_letter(j)}{i}', CellIsRule(operator = 'notEqual', formula = [vars.detalhe[key][j - 2]], fill = wrong))
