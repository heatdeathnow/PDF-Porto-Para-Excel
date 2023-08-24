import pdfplumber
from pdfplumber.page import Page
from pandas import DataFrame
import pandas as pd
from time import perf_counter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Color, Font, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from argparse import ArgumentParser


def crop_left_side(page: Page) -> Page:
    return page.within_bbox((0, 130, 690, 595))  # X0, Y0, X1, Y1

def crop_seguro(page: Page) -> Page:
    return page.within_bbox((0, 130, 70, 595))

def crop_dep(page: Page) -> Page:
    return page.within_bbox((60, 130, 100, 595))

def crop_nome(page: Page) -> Page:
    return page.within_bbox((100, 130, 250, 595))

def crop_regfunc(page: Page) -> Page:
    return page.within_bbox((250, 130, 300, 595))

def crop_idade(page: Page) -> Page:
    return page.within_bbox((300, 130, 330, 595))

def crop_parentesco(page: Page) -> Page:
    return page.within_bbox((330, 130, 380, 595))

def crop_plano(page: Page) -> Page:
    return page.within_bbox((380, 130, 540, 595))

def crop_mov(page: Page) -> Page:
    return page.within_bbox((540, 130, 610, 595))

def crop_situacao(page: Page) -> Page:
    return page.within_bbox((610, 130, 690, 595))

def crop_right_side(page: Page) -> Page:
    return page.within_bbox((690, 130, 842, 595))

def crop_headers(page: Page) -> Page:
    return page.within_bbox((0, 113, 842, 123))

def crop_qe_valores(page: Page) -> Page:
    return page.within_bbox((100, 138, 415, 215))

def crop_premio(page: Page) -> Page:
    return page.within_bbox((600, 138, 830, 215))

def crop_totalizador(page: Page) -> Page:
    return page.within_bbox((445, 238, 830, 345))

def crop_detalhamento(page: Page) -> Page:
    return page.within_bbox((0, 380, 842, 595))

# ---------

def update_sub(page: Page):
    global _sub
    cropped = page.within_bbox((0, 70, 375, 80))
    _sub = cropped.extract_text().replace('Subestipulante:', '').strip()

def get_cells(page: Page, rel_pos: list[int]) -> list:
    lines = page.extract_text_lines(return_chars = False)
    cells = [''] * len(rel_pos)
    flag = False

    for i, line in enumerate(lines):
        if i + 1 == len(lines) and not flag:  # Se é o último.
            index = rel_pos.index(round(line['top']))
            cells[index] = line['text']

        elif flag:
            flag = False
        
        elif abs(line['top'] - lines[i + 1]['top']) < 10:  # Se estiver muito pergunto verticalmente, considerar como duas linhas do mesmo texto
            index = rel_pos.index(round(line['top']))
            cells[index] = '\n'.join([line['text'], lines[i + 1]['text']])
            flag = True

        elif not flag:
            index = rel_pos.index(round(line['top']))
            cells[index] = line['text']

    return cells

def get_relative_positions(page: Page) -> list:  # Passar idade pois é mais fácil
    positions = []
    lines = page.extract_text_lines(return_chars = False)
    for line in lines:
        positions.append(round(line['top']))
    return positions

def get_info_from_leftside(page: Page) -> DataFrame:
    rel_pos = get_relative_positions(crop_idade(page))
    seguro = get_cells(crop_seguro(page), rel_pos)
    dep = get_cells(crop_dep(page), rel_pos)
    nome = get_cells(crop_nome(page), rel_pos)
    regfunc = get_cells(crop_regfunc(page), rel_pos)
    idade = get_cells(crop_idade(page), rel_pos)
    parentesco = get_cells(crop_parentesco(page), rel_pos)
    plano = get_cells(crop_plano(page), rel_pos)
    mov = get_cells(crop_mov(page), rel_pos)
    situacao = get_cells(crop_situacao(page), rel_pos)

    df = DataFrame({'Seguro': seguro,
                    'Dep': dep,
                    'Nome do Segurado': nome,
                    'Reg. Func.': regfunc,
                    'Idade': idade,
                    'Parentesco': parentesco,
                    'Plano': plano,
                    'Mov./Inic.Vig.': mov,
                    'Situação do Segurado': situacao})
    return df

def get_info_from_rightside(page: Page) -> DataFrame:
    rightside = crop_right_side(page)

    avoid_error = max(rightside.extract_text().count('Prêmio Base'), rightside.extract_text().count('Desc por Co-Part'))

    premio = [''] * avoid_error
    consultasq = [''] * avoid_error
    consultasv = [''] * avoid_error
    examesq = [''] * avoid_error
    examesv = [''] * avoid_error
    socorroq =[''] * avoid_error
    socorrov = [''] * avoid_error
    rataq = [''] * avoid_error
    ratav = [''] * avoid_error
    copart = [''] * avoid_error
    iof = [''] * avoid_error

    i = -1  # Prêmio Base vai ser sempre o primeiro valor.
    for line in rightside.extract_text_lines(return_chars = False):
        if 'Prêmio Base' in line['text']:
            premio[i + 1] = line['text'][12:].strip()
            i += 1

        elif 'CONSULTAS' in line['text']:
            end = line['text'].index(')')
            consultasq[i] = line['text'][11:end]  # O "(" sempre estará na 11ª posição.
            consultasv[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

        elif 'EXAMES' in line['text']:
            end = line['text'].index(')')
            examesq[i] = line['text'][8:end]  # O "(" sempre estará na 8ª posição.
            examesv[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

        elif 'PRONTO-SOCORRO' in line['text']:
            end = line['text'].index(')')
            socorroq[i] = line['text'][16:end]  # O "(" sempre estará na 16ª posição.
            socorrov[i] = line['text'][end + 1:].strip()  # Começando depois do ")" retire todos os espaços.

        elif 'Pro-Rata' in line['text']:
            end = line['text'].index(')')
            end_ = line['text'].index(')', end + 1)
            rataq[i] = line['text'][10:end]  # O "(" sempre estará na 16ª posição.
            ratav[i] = line['text'][end_ + 1:].strip()  # Começando depois do ")" retire todos os espaços.

        elif 'Desc por Co-Part' in line['text']:
            if copart[i] != '':  # Pode acontecer de faltar o prêmio por erro da Porto Seguro.
                i += 1
            copart[i] = line['text'][22:].strip()
        
        elif 'IOF' in line['text']:
            if iof[i] != '' and iof[i] != '0,00' and iof[i] != 0:
                pass  # Por algum motivo, os PDFs da Porto podem vir com dois IOFs, o segundo erroneamente zerado.
            else:
                iof[i] = line['text'][4:].strip()
        
        elif 'TOTAL' not in line['text'].upper():
            raise TypeError(f"Não foi possível ler a linha: \n{line['text']}")

    df = DataFrame({'Prêmio Base': premio,
                    'CONSULTAS (quantidade)': consultasq,
                    'CONSULTAS (valor)': consultasv,
                    'EXAMES (quantidade)': examesq,
                    'EXAMES (valor)': examesv,
                    'PRONTO-SOCORRO (quantidade)': socorroq,
                    'PRONTO-SOCORRO (valor)': socorrov,
                    'Pro-Rata (quantidade)': rataq,
                    'Pro-Rata (valor)': ratav,
                    'Desc por Co-Part. (-)': copart,
                    'IOF': iof})
    return df

def get_info_from_qe_valores(page: Page) -> DataFrame:
    cells = {}

    qev = crop_qe_valores(page)
    for i, row in enumerate(qev.extract_text_lines()):
        if i == 1: continue

        for j, value in enumerate(row['text'].split(' ')):
            if j % 2 == 0:
                cells[f'{get_column_letter(j + 2)}{i + 2}'] = int(value.replace('.', '').replace(',', '.').strip())  # B2:G2 & B4:G6
            else:
                cells[f'{get_column_letter(j + 2)}{i + 2}'] = float(value.replace('.', '').replace(',', '.').strip())
    return cells

def get_info_from_premio(page: Page) -> DataFrame:
    cells = {}
    premio = crop_premio(page)

    for i, row in enumerate(premio.extract_text_lines()):
        if i == 3: continue

        for j, value in enumerate(row['text'].split(' ')):
            cells[f'{get_column_letter(j + 2)}{i + 2}'] = float(value.replace('.', '').replace(',', '.').strip())
    return cells

def get_info_from_totalizador(page: Page) -> DataFrame:
    total = crop_totalizador(page)
    cells = {}

    for i, row in enumerate(total.extract_text_lines()):
        j = row['text'].index(':') + 2
        cells[f'B{i + 1}'] = float(row['text'][j:].replace('.', '').replace(',', '.').strip())
    return cells

def get_info_from_detalhamento(page: Page) -> DataFrame:
    detalhe = crop_detalhamento(page)
    cells = {}

    for row in detalhe.extract_text_lines():
        if 'Em atendimento à Lei' in row['text']: break
        l = 0
        for i in 'qwertyuiopasdfghjklçzxcvbnm':  # Isso descobre a posição da última letra na linha.
            j = row['text'].lower().rfind(i)
            if j > l: l = j

        hold = []
        for i, cell in enumerate(row['text'][l + 2:].split(' ')):
            if i == 3: continue

            if ',' in cell:
                hold.append(float(cell.replace('.', '').replace(',', '.').strip()))
            else:
                hold.append(int(cell.strip()))
        cells[row['text'][:l + 1]] = hold
    return cells

def get_info_from_page(page: Page) -> DataFrame:
    global _sub_has_changed
    if _sub == '' or _sub_has_changed:
        update_sub(page)
        _sub_has_changed = False
        print(f'\n\n--- Subestipulante atualmente em leitura: {_sub} ---\n')

    check = crop_headers(page)
    if len(check.extract_words()) == 0:
        _sub_has_changed = True
        return DataFrame({'Seguro': [],
                          'Dep': [],
                          'Nome do Segurado': [],
                          'Reg. Func.': [],
                          'Idade': [],
                          'Parentesco': [],
                          'Plano': [],
                          'Mov./Inic.Vig.': [],
                          'Situação do Segurado': [],
                          'Prêmio Base': [],
                          'CONSULTAS (quantidade)': [],
                          'CONSULTAS (valor)': [],
                          'EXAMES (quantidade)': [],
                          'EXAMES (valor)': [],
                          'PRONTO-SOCORRO (quantidade)': [],
                          'PRONTO-SOCORRO (valor)': [],
                          'Pro-Rata (quantidade)': [],
                          'Pro-Rata (valor)': [],
                          'Desc por Co-Part. (-)': [],
                          'IOF': []})
    
    else:
        dfl = get_info_from_leftside(page)
        dfr = get_info_from_rightside(page)
        dfs = DataFrame({'Subestipulante': [_sub for _ in range(len(dfl.index))]})
        return pd.concat([dfs, dfl, dfr], axis = 'columns')

def fix_types(df: DataFrame) -> DataFrame:
    for i in range(len(df.index)):
        if df.loc[i, 'Idade'] != '':
            df.loc[i, 'Idade'] = int(df.loc[i, 'Idade'])

        if df.loc[i, 'Prêmio Base'] != '':
            df.loc[i, 'Prêmio Base'] = float(df.loc[i, 'Prêmio Base'].replace('.', '').replace(',', '.'))

        if df.loc[i, 'CONSULTAS (quantidade)'] != '':
            df.loc[i, 'CONSULTAS (quantidade)'] = int(df.loc[i, 'CONSULTAS (quantidade)'])
        
        if df.loc[i, 'CONSULTAS (valor)'] != '':
            df.loc[i, 'CONSULTAS (valor)'] = float(df.loc[i, 'CONSULTAS (valor)'].replace('.', '').replace(',', '.'))
        
        if df.loc[i, 'EXAMES (quantidade)'] != '':
            df.loc[i, 'EXAMES (quantidade)'] = int(df.loc[i, 'EXAMES (quantidade)'])

        if df.loc[i, 'EXAMES (valor)'] != '':
            df.loc[i, 'EXAMES (valor)'] = float(df.loc[i, 'EXAMES (valor)'].replace('.', '').replace(',', '.'))

        if df.loc[i, 'PRONTO-SOCORRO (quantidade)'] != '':
            df.loc[i, 'PRONTO-SOCORRO (quantidade)'] = int(df.loc[i, 'PRONTO-SOCORRO (quantidade)'])

        if df.loc[i, 'PRONTO-SOCORRO (valor)'] != '':
            df.loc[i, 'PRONTO-SOCORRO (valor)'] = float(df.loc[i, 'PRONTO-SOCORRO (valor)'].replace('.', '').replace(',', '.'))

        if df.loc[i, 'Pro-Rata (quantidade)'] != '':
            df.loc[i, 'Pro-Rata (quantidade)'] = int(df.loc[i, 'Pro-Rata (quantidade)'])
        
        if df.loc[i, 'Pro-Rata (valor)'] != '':
            df.loc[i, 'Pro-Rata (valor)'] = float(df.loc[i, 'Pro-Rata (valor)'].replace('.', '').replace(',', '.'))
        
        if df.loc[i, 'Desc por Co-Part. (-)'] != '':
            df.loc[i, 'Desc por Co-Part. (-)'] = float(df.loc[i, 'Desc por Co-Part. (-)'].replace('.', '').replace(',', '.'))
        
        if df.loc[i, 'IOF'] != '':
            df.loc[i, 'IOF'] = float(df.loc[i, 'IOF'].replace('.', '').replace(',', '.'))

def add_cols_extracted(ws: Worksheet) -> None:
    ws['V1'] = 'Coparticipação'
    ws['W1'] = 'TOTAL DO DEP.'
    ws['X1'] = 'Prêmio Real'
    for i in range(2, ws.max_row + 1):
        ws[f'V{i}'] = f'=ROUND($M{i} + $O{i} + $Q{i}, 2)'
        ws[f'W{i}'] = f'=ROUND($K{i} + $S{i} - $T{i} + $U{i} + $V{i}, 2)'
        ws[f'X{i}'] = f'=ROUND($K{i} + $U{i}, 2)'
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

def add_cols_qe_valor(ws: Worksheet, m: int) -> None:
    ws['A2'] = 'Inclusões'
    ws['A3'] = 'Reativações'
    ws['A4'] = 'Alterações'
    ws['A5'] = 'Exclusões'
    ws['A6'] = 'Atual'

    ws['C1'] = 'Titulares'
    # Quantidade de titulares incluídos
    ws['B2'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "I", \'Informação extraída\'!$G$2:$G${m}, "Titular")'
    # Soma dos totais consumidos pelos beneficiários titulares incluídos. (total dep.)
    ws['C2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "I", \'Informação extraída\'!$G$2:$G${m}, "Titular"), 2)'

    # Com alterações
    ws['B4'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "Titular")'
    ws['C4'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "Titular"), 2)'

    # Excluídos
    ws['B5'] = f'=COUNTIFS(\'Informação extraída\'!$I$2:$I${m}, "Exc\\", \'Informação extraída\'!$G$2:$G${m}, "Titular")'
    ws['C5'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$I$2:$I${m}, "Exc\\", \'Informação extraída\'!$G$2:$G${m}, "Titular"), 2)'

    # Beneficiários ativos (todos menos os excluídos)
    ws['B6'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "", \'Informação extraída\'!$G$2:$G${m}, "Titular") + SUM(B2:B4)'
    ws['C6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "", \'Informação extraída\'!$G$2:$G${m}, "Titular") + SUM(C2:C4), 2)'

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    ws['E1'] = 'Dependentes'
    # Inclusões
    ws['D2'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "I", \'Informação extraída\'!$G$2:$G${m}, "<>Titular")'
    ws['E2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "I", \'Informação extraída\'!$G$2:$G${m}, "<>Titular"), 2)'

    # Alterações
    ws['D4'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "<>Titular")'
    ws['E4'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "<>Titular"), 2)'

    # Exclusões
    ws['D5'] = f'=COUNTIFS(\'Informação extraída\'!$I$2:$I${m}, "Exc\\", \'Informação extraída\'!$G$2:$G${m}, "<>Titular")'
    ws['E5'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$I$2:$I${m}, "Exc\\", \'Informação extraída\'!$G$2:$G${m}, "<>Titular"), 2)'

    # Ativos
    ws['D6'] = f'=COUNTIFS(\'Informação extraída\'!$J$2:$J${m}, "", \'Informação extraída\'!$G$2:$G${m}, "<>Titular") + SUM(D2:D4)'
    ws['E6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!W$2:$W${m}, \'Informação extraída\'!$J$2:$J${m}, "", \'Informação extraída\'!$G$2:$G${m}, "<>Titular") + SUM(E2:E4), 2)'

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    ws['G1'] = 'Totais'
    # Inclusões
    ws['F2'] = f'=ROUND(SUM(B2, D2), 2)'  # quantidade
    ws['G2'] = f'=ROUND(SUM(C2, E2), 2)'  # valor

    # Alterações
    ws['F4'] = f'=ROUND(SUM(B4, D4), 2)'  # quantidade
    ws['G4'] = f'=ROUND(SUM(C4, E4), 2)'  # valor

    # Exclusões
    ws['F5'] = f'=ROUND(SUM(B5, D5), 2)'  # quantidade
    ws['G5'] = f'=ROUND(SUM(C5, E5), 2)'  # valor

    # Ativos
    ws['F6'] = f'=ROUND(SUM(B6, D6), 2)'  # quantidade
    ws['G6'] = f'=ROUND(SUM(C6, E6), 2)'  # valor

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    # A Porto é bugada, por algum motivo alterações é mostrado como soma do "prêmio real", mas é considerado como soma do "total dep." no cálculo dos ativos.
    ws['I1'] = 'Alterações como aparecerá no PDF:'
    ws['I2'] = 'Titulares'
    ws['J2'] = 'Dependentes'
    ws['I3'] = f'=ROUND(SUMIFS(\'Informação extraída\'!K$2:$K${m}, \'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "Titular"), 2)'
    ws['J3'] = f'=ROUND(SUMIFS(\'Informação extraída\'!K$2:$K${m}, \'Informação extraída\'!$J$2:$J${m}, "A", \'Informação extraída\'!$G$2:$G${m}, "<>Titular"), 2)'

def add_premio(ws: Worksheet, m: int) -> None:
    ws['A2'] = 'Prêmio Real(=)'
    ws['A3'] = 'Co-Participação(+)'
    ws['A4'] = 'Desc. Co=Participação(-)'
    ws['A5'] = 'Desp.Não Cobertas (+)'
    ws['A6'] = 'Totais'

    ws['B1'] = 'Titulares'
    ws['B2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!$X$2:$X${m}, \'Informação extraída\'!$G$2:$G${m}, "Titular", \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\") + SUMIF(\'Informação extraída\'!$G$2:$G${m}, "Titular", \'Informação extraída\'!$S$2:$S${m}), 2)'
    ws['B3'] = f'=ROUND(SUMIF(\'Informação extraída\'!$G$2:$G${m}, "Titular", \'Informação extraída\'!$V2:$V${m}), 2)'
    ws['B4'] = f'= - ROUND(SUMIFS(\'Informação extraída\'!$T$2:$T${m}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\", \'Informação extraída\'!$G2:$G{m}, "Titular"), 2)'
    ws['B6'] = '=ROUND(SUM(B2:B5), 2)'

    ws['C1'] = 'Dependentes'
    ws['C2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!$X$2:$X${m}, \'Informação extraída\'!$G$2:$G${m}, "<>Titular", \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\") + SUMIF(\'Informação extraída\'!$G$2:$G${m}, "<>Titular", \'Informação extraída\'!$S$2:$S${m}), 2)'
    ws['C3'] = f'=ROUND(SUMIF(\'Informação extraída\'!$G$2:$G${m}, "<>Titular", \'Informação extraída\'!$V2:$V${m}), 2)'
    ws['C4'] = f'= - ROUND(SUMIFS(\'Informação extraída\'!$T$2:$T${m}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\", \'Informação extraída\'!$G2:$G{m}, "<>Titular"), 2)'
    ws['C6'] = '=ROUND(SUM(C2:C5), 2)'

    ws['D1'] = 'Total'
    ws['D2'] = '=ROUND(SUM(B2:C2), 2)'
    ws['D3'] = '=ROUND(SUM(B3:C3), 2)'
    ws['D4'] = '=ROUND(SUM(B4:C4), 2)'
    ws['D6'] = '=ROUND(SUM(B6:C6), 2)'

def add_totalizador(ws: Worksheet, m: int) -> None:
    ws['A1'] = 'Prêmio Líquido'
    ws['A2'] = 'Taxa de Inscrição'
    ws['A3'] = 'Prêmio Pró-Rata'
    ws['A4'] = 'Bônus'
    ws['A5'] = 'Base I.O.F. (2,38%)'
    ws['A6'] = 'I.O.F'
    ws['A7'] = 'Despesas Não Cobertas'
    ws['A8'] = 'Total'

    ws['B1'] = '=ROUND(B5 - B3, 2)'
    ws['B3'] = f'=ROUND(SUM(\'Informação extraída\'!$S$2:$S${m}), 2)'
    ws['B5'] = '=ROUND(\'Prêmio\'!$D$6 - B6, 2)'
    ws['B6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!$U$2:$U${m}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\"), 2)'
    ws['B8'] = '=ROUND(B1 + B3 + B6, 2)'

def add_detalhamento(ws: Worksheet, base: Worksheet) -> None:
    m = base.max_row
    planos = []
    for cell in base.iter_rows(min_row = 2):
        if cell[7].value not in planos:
            planos.append(cell[7].value)
    
    ws['A1'] = 'Planos'
    ws['B1'] = 'Total Segurados'
    ws['C1'] = 'Total Dep. Especiais'
    ws['D1'] = 'Despesas Não Coberta'
    ws['E1'] = 'Total de Prêmio\Custo Total'
    ws['F1'] = 'Pró-Rata'

    for i, plano in enumerate(planos, 2):
        ws[f'A{i}'] = plano
        ws[f'B{i}'] = f'=COUNTIFS(\'Informação extraída\'!$H$2:$H${m}, $A{i}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\")'
        ws[f'E{i}'] = f'=ROUND(SUMIFS(\'Informação extraída\'!$W$2:$W${m}, \'Informação extraída\'!$H$2:$H${m}, $A{i}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\") + SUMIFS(\'Informação extraída\'!$V$2:$V${m}, \'Informação extraída\'!$H$2:$H${m}, $A{i}, \'Informação extraída\'!$I$2:$I${m}, "Exc\\"), 2)'
        ws[f'F{i}'] = f'=ROUND(SUMIFS(\'Informação extraída\'!$S$2:$S${m}, \'Informação extraída\'!$H$2:$H${m}, $A{i}, \'Informação extraída\'!$I$2:$I${m}, "<>Exc\\"), 2)'

    ws[f'E{len(planos) + 2}'] = f'=ROUND(SUM(E2:E{len(planos) + 1}), 2)' # 1 dos cabeçários, 1 para ficar embaixo do último.
    ws[f'F{len(planos) + 2}'] = f'=ROUND(SUM(F2:F{len(planos) + 1}), 2)'

def format_sheets(wb: Workbook) -> None:
    blue = Color('00002060')
    white = Color('00FFFFFF')
    black = Color('00000000')
    side = Side(color = black, style = 'thin')
    borders = Border(side, side, side, side)
    blue_fill = PatternFill(start_color = blue, end_color = blue, fill_type = 'solid')
    header_font = Font(color = white, bold = True)

    ws = wb['Informação extraída']
    ws.sheet_view.showGridLines = False
    for i in range(1, ws.max_column + 1):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, ws.max_row + 1):
            ws[f'J{j}'].number_format = _nformat
            ws[f'L{j}'].number_format = _nformat
            ws[f'N{j}'].number_format = _nformat
            ws[f'P{j}'].number_format = _nformat
            ws[f'R{j}'].number_format = _nformat
            ws[f'S{j}'].number_format = _nformat
            ws[f'T{j}'].number_format = _nformat
            ws[f'U{j}'].number_format = _nformat
            ws[f'V{j}'].number_format = _nformat
            ws[f'W{j}'].number_format = _nformat
            ws[f'{get_column_letter(i)}{j}'].border = borders

    ws = wb['Quantidades e valores']
    ws.sheet_view.showGridLines = False
    for i in range(1, 8):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, 7):
            ws[f'{get_column_letter(i)}{j}'].border = borders
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font
            ws[f'C{j}'].number_format = _nformat
            ws[f'E{j}'].number_format = _nformat
            ws[f'G{j}'].number_format = _nformat
    ws[f'I3'].number_format = _nformat
    ws[f'J3'].number_format = _nformat
    ws[f'I3'].border = borders
    ws[f'J3'].border = borders
    ws[f'I1'].border = borders
    ws[f'I2'].border = borders
    ws[f'J2'].border = borders
    ws[f'I2'].fill = blue_fill
    ws[f'I1'].fill = blue_fill
    ws[f'J2'].fill = blue_fill
    ws[f'I2'].font = header_font
    ws[f'I1'].font = header_font
    ws[f'J2'].font = header_font
    ws.merge_cells('I1:J1')

    ws = wb['Prêmio']
    ws.sheet_view.showGridLines = False
    for i in range(1, 5):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, 7):
            ws[f'{get_column_letter(i)}{j}'].border = borders
            if i == 1:
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font
            else: 
                ws[f'{get_column_letter(i)}{j}'].number_format = _nformat

    ws = wb['Totalizador']
    ws.sheet_view.showGridLines = False
    for i in range(1, 3):
        if i != 2:
            ws[f'{get_column_letter(i)}1'].fill = blue_fill
            ws[f'{get_column_letter(i)}1'].font = header_font
        else:
            ws[f'{get_column_letter(i)}1'].number_format = _nformat

        ws[f'{get_column_letter(i)}1'].border = borders
        for j in range(2, 9):
            ws[f'{get_column_letter(i)}{j}'].border = borders
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font
            else: 
                ws[f'{get_column_letter(i)}{j}'].number_format = _nformat

    ws = wb['Detalhamento']
    ws.sheet_view.showGridLines = False
    for i in range(1, 7):
        ws[f'{get_column_letter(i)}1'].fill = blue_fill
        ws[f'{get_column_letter(i)}1'].font = header_font
        ws[f'{get_column_letter(i)}1'].border = borders

        for j in range(2, ws.max_row):
            ws[f'{get_column_letter(i)}{j}'].border = borders
            if i == 1: 
                ws[f'{get_column_letter(i)}{j}'].fill = blue_fill
                ws[f'{get_column_letter(i)}{j}'].font = header_font
            elif i != 2: 
                ws[f'{get_column_letter(i)}{j}'].number_format = _nformat
    ws[f'E{ws.max_row}'].number_format = _nformat
    ws[f'F{ws.max_row}'].number_format = _nformat
    ws[f'E{ws.max_row}'].border = borders
    ws[f'F{ws.max_row}'].border = borders

def check_values(wb: Workbook) -> None:
    global _qev
    red = Color('00FF0000')
    green = Color('00008000')
    wrong = PatternFill(start_color = red, end_color = red, fill_type = 'solid')
    right = PatternFill(start_color = green, end_color = green, fill_type = 'solid')

    ws = wb['Quantidades e valores']
    for key in _qev.keys():
        ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [_qev[key]], fill = right))
        ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [_qev[key]], fill = wrong))
    ws.conditional_formatting.add('I3', CellIsRule(operator = 'equal', formula = [_qev['C4']], fill = right))
    ws.conditional_formatting.add('I3', CellIsRule(operator = 'notEqual', formula = [_qev['C4']], fill = wrong))
    ws.conditional_formatting.add('J3', CellIsRule(operator = 'equal', formula = [_qev['E4']], fill = right))
    ws.conditional_formatting.add('J3', CellIsRule(operator = 'notEqual', formula = [_qev['E4']], fill = wrong))

    ws = wb['Prêmio']
    for key in _premio.keys():
        if '4' not in key:
            ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [_premio[key]], fill = right))
            ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [_premio[key]], fill = wrong))
        else:
            ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [-_premio[key]], fill = right))
            ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [-_premio[key]], fill = wrong))

    ws = wb['Totalizador']
    for key in _total.keys():
        ws.conditional_formatting.add(key, CellIsRule(operator = 'equal', formula = [_total[key]], fill = right))
        ws.conditional_formatting.add(key, CellIsRule(operator = 'notEqual', formula = [_total[key]], fill = wrong))
    
    ws = wb['Detalhamento']
    for i in range(2, ws.max_row):
        for j in range(2, 6):
            key = str(ws[f'A{i}'].value)
            key = key.replace('\n', ' ')
            ws.conditional_formatting.add(f'{get_column_letter(j)}{i}', CellIsRule(operator = 'equal', formula = [_detalhe[key][j - 2]], fill = right))
            ws.conditional_formatting.add(f'{get_column_letter(j)}{i}', CellIsRule(operator = 'notEqual', formula = [_detalhe[key][j - 2]], fill = wrong))


_sub = ''
_sub_has_changed = False
_nformat = 'R$* #,##0.00;-R$* #,##0.00;-;'
if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('input')
    args = parser.parse_args()
    file = args.input

    if file[-4:] != '.pdf':
        raise TypeError('Apenas arquivos .PDF são permitidos.')
    output = file.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')

    runtime = perf_counter()
    with pdfplumber.open(file) as pdf:
        amount = len(pdf.pages)

        _qev = get_info_from_qe_valores(pdf.pages[0])
        _premio = get_info_from_premio(pdf.pages[0])
        _total = get_info_from_totalizador(pdf.pages[0])
        _detalhe = get_info_from_detalhamento(pdf.pages[0])
        for i in range(1, amount - 1):  # Ignora a primeira e última página.
            start_time = perf_counter()
            print(f'Extraindo informação da página {i + 1:02}ª... ', end = '')
            df = get_info_from_page(pdf.pages[i])
            print(f'informação extraída em {perf_counter() - start_time:.2f} segundos.')

            if i == 1:
                baseframe = df.copy()
            else:
                baseframe = pd.concat([baseframe, df], ignore_index = True)

    fix_types(baseframe)

    start_time = perf_counter()
    print(f'Salvando a informação no disco... ', end = '')
    writer = pd.ExcelWriter(output, 'openpyxl')
    baseframe.to_excel(writer, 'Informação extraída', index = False)
    print(f'Informação salva em {perf_counter() - start_time:.2f} segundos.')

    start_time = perf_counter()
    print(f'Adicionando fórmulas do Excel... ', end = '')
    add_cols_extracted(writer.sheets['Informação extraída'])  # As duas colunas na primeira planilha
    writer.book.create_sheet('Quantidades e valores')
    add_cols_qe_valor(writer.sheets['Quantidades e valores'], writer.sheets['Informação extraída'].max_row)  # Segunda planilha com informações de inclusões, reativações, alterações, exclusões
    writer.book.create_sheet('Prêmio')
    add_premio(writer.sheets['Prêmio'], writer.sheets['Informação extraída'].max_row)  # Terceira planilha com informações de coparticipação etc
    writer.book.create_sheet('Totalizador')
    add_totalizador(writer.sheets['Totalizador'], writer.sheets['Informação extraída'].max_row)  # Quarta planilha com segurados e prêmios por plano
    writer.book.create_sheet('Detalhamento')
    add_detalhamento(writer.sheets['Detalhamento'], writer.sheets['Informação extraída'])
    print(f'fórmulas adicionadas em {perf_counter() - start_time:.2f} segundos.')

    start_time = perf_counter()
    print(f'Formatando planilhas... ', end = '')
    writer.sheets.update(dict((ws.title, ws) for ws in writer.book.worksheets))  # ??? Necessário
    format_sheets(writer.book)
    check_values(writer.book)
    print(f'planilhas formatadas em {perf_counter() - start_time:.2f} segundos.')

    writer.close()
    print(f'Tempo total de execução: {perf_counter() - runtime:.2f} segundos.')
