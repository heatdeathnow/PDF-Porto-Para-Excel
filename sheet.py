from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from vars import *


def add_cols_extracted(ws: Worksheet) -> None:
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

    ws[f'{COP}1'] = 'Coparticipação'
    ws[f'{TDE}1'] = 'TOTAL DO DEP.'
    ws[f'{PRR}1'] = 'Prêmio Real'

    for i in range(2, ws.max_row + 1):
        ws[f'{COP}{i}'] = f'=ROUND(${COV}{i} + ${EXV}{i} + ${PRV}{i}, 2)'
        ws[f'{TDE}{i}'] = f'=ROUND(${PRE}{i} + ${RAV}{i} - ${DES}{i} - {BON}{i} - {ACE}{i} + ${IOF}{i} + ${COP}{i}, 2)'
        ws[f'{PRR}{i}'] = f'=ROUND(${PRE}{i} - {BON}{i} - {ACE}{i} + ${IOF}{i}, 2)'

def add_cols_qe_valor(ws: Worksheet, m: int) -> None:
    ws['A2'] = 'Inclusões'
    ws['A3'] = 'Reativações'
    ws['A4'] = 'Alterações'
    ws['A5'] = 'Exclusões'
    ws['A6'] = 'Atual'

    ws['C1'] = 'Titulares'
    # Quantidade de titulares incluídos
    ws['B2'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "I", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular")'
    # Soma dos totais consumidos pelos beneficiários titulares incluídos. (total dep.)
    ws['C2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "I", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular"), 2)'

    # Com alterações
    ws['B4'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular")'
    ws['C4'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular"), 2)'

    # Excluídos
    ws['B5'] = f'=COUNTIFS(\'Informação extraída\'!${MOV}$2:${MOV}${m}, "Exc\\", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular")'
    ws['C5'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "Exc\\", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular"), 2)'

    # Beneficiários ativos (todos menos os excluídos)
    ws['B6'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular") + SUM(B2:B4)'
    ws['C6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular") + SUM(C2:C4), 2)'

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    ws['E1'] = 'Dependentes'
    # Inclusões
    ws['D2'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "I", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular")'
    ws['E2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "I", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular"), 2)'

    # Alterações
    ws['D4'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular")'
    ws['E4'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular"), 2)'

    # Exclusões
    ws['D5'] = f'=COUNTIFS(\'Informação extraída\'!${MOV}$2:${MOV}${m}, "Exc\\", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular")'
    ws['E5'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "Exc\\", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular"), 2)'

    # Ativos
    ws['D6'] = f'=COUNTIFS(\'Informação extraída\'!${SIT}$2:${SIT}${m}, "", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular") + SUM(D2:D4)'
    ws['E6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{TDE}$2:${TDE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular") + SUM(E2:E4), 2)'

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    ws['G1'] = 'Totais'
    # Inclusões
    ws['F2'] = f'=ROUND(SUM(B2, D2), 2)'  # quantidade
    ws['G2'] = f'=ROUND(SUM(C2, E2), 2)'  # valor

    # Alterações
    ws['F4'] = f'=ROUND(SUM(B4, D4), 2)'  # quantidade
    ws['G4'] = f'=ROUND(SUM(C4, E4), 2)'  # valor

    # Exclusões
    ws['F5'] = f'=ROUND(SUM(B5, D5), 2)'  # quantidade
    ws['G5'] = f'=ROUND(SUM(C5, E5), 2)'  # valor

    # Ativos
    ws['F6'] = f'=ROUND(SUM(B6, D6), 2)'  # quantidade
    ws['G6'] = f'=ROUND(SUM(C6, E6), 2)'  # valor

    # ----------------------------------------------------------------------------------------------------------------------------------------------- #
    # A Porto é bugada, por algum motivo alterações é mostrado como soma do "prêmio real", mas é considerado como soma do "total dep." no cálculo dos ativos.
    ws['I1'] = 'Alterações como aparecerá no PDF:'
    ws['I2'] = 'Titulares'
    ws['J2'] = 'Dependentes'
    ws['I3'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{PRE}$2:${PRE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular"), 2)'
    ws['J3'] = f'=ROUND(SUMIFS(\'Informação extraída\'!{PRE}$2:${PRE}${m}, \'Informação extraída\'!${SIT}$2:${SIT}${m}, "A", \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular"), 2)'

def add_premio(ws: Worksheet, m: int) -> None:
    ws['A2'] = 'Prêmio Real(=)'
    ws['A3'] = 'Co-Participação(+)'
    ws['A4'] = 'Desc. Co=Participação(-)'
    ws['A5'] = 'Desp.Não Cobertas (+)'
    ws['A6'] = 'Totais'

    ws['B1'] = 'Titulares'
    ws['B2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!${PRR}$2:${PRR}${m}, \'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular", \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\") + SUMIF(\'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular", \'Informação extraída\'!${RAV}$2:${RAV}${m}), 2)'
    ws['B3'] = f'=ROUND(SUMIF(\'Informação extraída\'!${PAR}$2:${PAR}${m}, "Titular", \'Informação extraída\'!${COP}2:${COP}${m}), 2)'
    ws['B4'] = f'= - ROUND(SUMIFS(\'Informação extraída\'!${DES}$2:${DES}${m}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\", \'Informação extraída\'!${PAR}2:${PAR}{m}, "Titular"), 2)'
    ws['B6'] = '=ROUND(SUM(B2:B5), 2)'

    ws['C1'] = 'Dependentes'
    ws['C2'] = f'=ROUND(SUMIFS(\'Informação extraída\'!${PRR}$2:${PRR}${m}, \'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular", \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\") + SUMIF(\'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular", \'Informação extraída\'!${RAV}$2:${RAV}${m}), 2)'
    ws['C3'] = f'=ROUND(SUMIF(\'Informação extraída\'!${PAR}$2:${PAR}${m}, "<>Titular", \'Informação extraída\'!${COP}2:${COP}${m}), 2)'
    ws['C4'] = f'= - ROUND(SUMIFS(\'Informação extraída\'!${DES}$2:${DES}${m}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\", \'Informação extraída\'!${PAR}2:${PAR}{m}, "<>Titular"), 2)'
    ws['C6'] = '=ROUND(SUM(C2:C5), 2)'

    ws['D1'] = 'Total'
    ws['D2'] = '=ROUND(SUM(B2:C2), 2)'
    ws['D3'] = '=ROUND(SUM(B3:C3), 2)'
    ws['D4'] = '=ROUND(SUM(B4:C4), 2)'
    ws['D6'] = '=ROUND(SUM(B6:C6), 2)'

def add_totalizador(ws: Worksheet, m: int) -> None:
    ws['A1'] = 'Prêmio Líquido'
    ws['A2'] = 'Taxa de Inscrição'
    ws['A3'] = 'Prêmio Pró-Rata'
    ws['A4'] = 'Bônus'
    ws['A5'] = 'Base I.O.F. (2,38%)'
    ws['A6'] = 'I.O.F'
    ws['A7'] = 'Despesas Não Cobertas'
    ws['A8'] = 'Total'

    ws['B1'] = '=ROUND(B5 - B3 + B4, 2)'
    ws['B3'] = f'=ROUND(SUM(\'Informação extraída\'!${RAV}$2:${RAV}${m}), 2)'
    ws['B4'] = f'=ROUND(SUM(\'Informação extraída\'!${BON}$2:${BON}${m}), 2)'
    ws['B5'] = '=ROUND(\'Prêmio\'!$D$6 - B6, 2)'
    ws['B6'] = f'=ROUND(SUMIFS(\'Informação extraída\'!${IOF}$2:${IOF}${m}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\"), 2)'
    ws['B8'] = '=ROUND(B1 + B3 - B4 + B6, 2)'

def add_detalhamento(ws: Worksheet, base: Worksheet) -> None:
    m = base.max_row
    planos = []
    
    for cell in base.iter_rows(min_row = 2):
        if cell[7].value not in planos:
            planos.append(cell[7].value)
    
    ws['A1'] = 'Planos'
    ws['B1'] = 'Total Segurados'
    ws['C1'] = 'Total Dep. Especiais'
    ws['D1'] = 'Despesas Não Coberta'
    ws['E1'] = 'Total de Prêmio\Custo Total'
    ws['F1'] = 'Pró-Rata'

    for i, plano in enumerate(planos, 2):
        ws[f'A{i}'] = plano
        ws[f'B{i}'] = f'=COUNTIFS(\'Informação extraída\'!${PLA}$2:${PLA}${m}, $A{i}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\")'
        ws[f'E{i}'] = f'=ROUND(SUMIFS(\'Informação extraída\'!${TDE}$2:${TDE}${m}, \'Informação extraída\'!${PLA}$2:${PLA}${m}, $A{i}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\") + SUMIFS(\'Informação extraída\'!${COP}$2:${COP}${m}, \'Informação extraída\'!${PLA}$2:${PLA}${m}, $A{i}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "Exc\\"), 2)'
        ws[f'F{i}'] = f'=ROUND(SUMIFS(\'Informação extraída\'!${RAV}$2:${RAV}${m}, \'Informação extraída\'!${PLA}$2:${PLA}${m}, $A{i}, \'Informação extraída\'!${MOV}$2:${MOV}${m}, "<>Exc\\"), 2)'

    ws[f'E{len(planos) + 2}'] = f'=ROUND(SUM(E2:E{len(planos) + 1}), 2)' # 1 dos cabeçários, 1 para ficar embaixo do último.
    ws[f'F{len(planos) + 2}'] = f'=ROUND(SUM(F2:F{len(planos) + 1}), 2)'
